import io
import os
from datetime import datetime

import pandas as pd
import requests
import streamlit as st
import plotly.express as px


# ============================================================
# CONFIGURACIÓN
# ============================================================

st.set_page_config(
    page_title="Resumen del Mes",
    page_icon="📊",
    layout="wide",
)

API_BASE_URL = os.getenv(
    "API_BASE_URL",
    "https://app.granbodega.com.mx",
).rstrip("/")


# ============================================================
# CONSTANTES
# ============================================================

DICC_MESES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}

COLOR_MARCA = "#FF5100"


# ============================================================
# CSS GENERAL Y FUENTE DE ÍCONOS
# ============================================================

st.markdown(
    """
    <!-- Carga de Material Symbols de Google -->
    <link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:opsz,wght,FILL,GRAD@24,400,0,0" />

    <style>

    /* ========================================================
       CONTENEDOR PRINCIPAL
       ======================================================== */

    .block-container {
        padding-top: 1rem !important;
        padding-bottom: 3rem !important;
    }


    /* ========================================================
       TARJETAS KPI
       ======================================================== */

    .kpi-card {
        background-color: #ffffff;
        border: 1px solid #e2e8f0;
        border-left: 4px solid #FF5100;
        border-radius: 10px;
        padding: 12px 10px;
        height: 125px;
        display: flex;
        flex-direction: column;
        justify-content: space-between;
        box-sizing: border-box;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
        transition: all 0.25s ease;
        overflow: hidden;
    }

    .kpi-top {
        display: flex;
        flex-direction: column;
        gap: 2px;
    }

    .kpi-header {
        display: flex;
        align-items: center;
        gap: 6px;
        font-size: 14px;
        font-weight: 700;
        color: #64748b;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }

    .kpi-icon {
        font-family: 'Material Symbols Outlined' !important;
        font-size: 18px !important;
        line-height: 1 !important;
        display: inline-block;
        vertical-align: middle;
    }

    .kpi-value {
        font-size: 22px;
        font-weight: 800;
        color: #334FB5;
        line-height: 1.2;
        margin-top: 2px;
    }

    .kpi-label {
        font-size: 12px;
        color: #64748b;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }

    .kpi-badge-positive {
        display: inline-block;
        background-color: #dcfce7;
        color: #166534;
        font-size: 10px;
        font-weight: 700;
        padding: 2px 6px;
        border-radius: 4px;
        margin-top: 4px;
        width: fit-content;
    }

    .kpi-badge-negative {
        display: inline-block;
        background-color: #fee2e2;
        color: #991b1b;
        font-size: 10px;
        font-weight: 700;
        padding: 2px 6px;
        border-radius: 4px;
        margin-top: 4px;
        width: fit-content;
    }

    .kpi-badge-info {
        display: inline-block;
        background-color: #f1f5f9;
        color: #475569;
        font-size: 10px;
        font-weight: 600;
        padding: 2px 6px;
        border-radius: 4px;
        margin-top: 4px;
        width: fit-content;
    }


    /* ========================================================
       FILTROS
       ======================================================== */

    .filter-title {
        font-size: 18px;
        font-weight: 800;
        color: #334fb5;
        margin-top: 8px;
        margin-bottom: 10px;
    }


    /* ========================================================
       BOTONES
       ======================================================== */

    div.stButton > button[kind="primary"] {
        background: linear-gradient(
            135deg,
            #ff5100 0%,
            #e04700 100%
        ) !important;
        color: white !important;
        font-weight: 700 !important;
        border-radius: 8px !important;
        border: none !important;
        box-shadow: 0 4px 12px rgba(255, 81, 0, 0.25) !important;
    }

    div.stButton > button[kind="primary"]:hover {
        transform: translateY(-1px) !important;
        box-shadow: 0 6px 16px rgba(255, 81, 0, 0.35) !important;
    }


    /* ========================================================
       SELECTORES
       ======================================================== */

    div[data-baseweb="select"] > div {
        border-radius: 8px !important;
        border-color: #cbd5e1 !important;
    }


    /* ========================================================
       TABLAS HTML PROPIAS (render_tabla_html)
       ======================================================== */

    .tabla-html-wrapper {
        overflow-x: auto;
        border-radius: 8px;
        border: 1px solid #e2e8f0;
    }

    .tabla-html-wrapper table {
        border-collapse: collapse;
        width: 100%;
        font-size: 13px;
    }

    .tabla-html-wrapper thead th {
        background-color: #FF5100 !important;
        color: #FFFFFF !important;
        font-weight: 700 !important;
        text-align: left;
        padding: 10px 12px;
        position: sticky;
        top: 0;
    }

    .tabla-html-wrapper tbody td {
        padding: 8px 12px;
        border-bottom: 1px solid #e2e8f0;
        color: #0f172a;
    }

    .tabla-html-wrapper tbody tr:nth-child(even) {
        background-color: #f8fafc;
    }

    .tabla-html-wrapper tbody tr:hover {
        background-color: #fff1e8;
    }

    .section-header-container {
    background: linear-gradient(135deg, #ffffff 0%, #f0f7ff 100%);
    border: 1px solid #cbd5e1;
    border-left: 5px solid #334FB5;
    border-radius: 10px;
    padding: 12px 20px;
    margin-top: 10px;
    margin-bottom: 20px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    box-shadow: 0 2px 4px rgba(0, 0, 0, 0.03);
    }

    .section-header-left {
        display: flex;
        align-items: center;
        gap: 14px;
    }

    .section-header-icon {
        background: rgba(0, 51, 102, 0.08);
        border-radius: 8px;
        width: 40px;
        height: 40px;
        display: flex;
        align-items: center;
        justify-content: center;
        flex-shrink: 0;
    }

    .section-title {
        font-size: 26px;
        font-weight: 800;
        color: #334FB5;
        margin: 0;
        line-height: 1.2;
        letter-spacing: -0.3px;
        text-transform: uppercase;
    }

    .section-subtitle {
        font-size: 12px;
        font-weight: 500;
        color: #475569;
        margin-top: 2px;
    }

    .custom-divider {
        height: 1px;
        background: linear-gradient(90deg, #cbd5e1 0%, rgba(203, 213, 225, 0.2) 100%);
        margin-top: 25px;
        margin-bottom: 25px;
        border: none;
    }

    </style>
    """,
    unsafe_allow_html=True,
)
def render_section_header(icon, titulo, subtitulo=""):
    header_html = f"""
    <div class="section-header-container">
        <div class="section-header-left">
            <div class="section-header-icon">
                <span class="material-symbols-outlined">{icon}</span>
            </div>
            <div>
                <div class="section-title">{titulo}</div>
                {f'<div class="section-subtitle">{subtitulo}</div>' if subtitulo else ''}
            </div>
        </div>
    </div>
    """
    st.markdown(header_html, unsafe_allow_html=True)

def render_divider():
    st.markdown('<div class="custom-divider"></div>', unsafe_allow_html=True)

# ============================================================
# API
# ============================================================
render_section_header("analytics", "Resumen del Mes", "Análisis acumulado YTD, ventas y desempeño por cliente, sucursal y categoría")

@st.cache_data(ttl=300)
def obtener_datos(
    endpoint: str,
    params: dict | None = None,
):
    url = f"{API_BASE_URL}/{endpoint}"

    try:
        response = requests.get(
            url,
            params=params or {},
            timeout=260,
        )

        if response.status_code != 200:
            st.error(
                f"Error HTTP {response.status_code} "
                f"consultando '{endpoint}': "
                f"{response.text}"
            )
            return None

        res = response.json()

        if res.get("status") != "success":
            st.error(
                f"El endpoint '{endpoint}' "
                f"regresó un error: "
                f"{res.get('message', 'Error desconocido')}"
            )
            return None

        return res.get("data")

    except requests.exceptions.RequestException as e:
        st.error(
            f"No se pudo conectar con el API: {e}"
        )
        return None

    except Exception as e:
        st.error(
            f"Error consultando '{endpoint}': {e}"
        )
        return None


# ============================================================
# CATÁLOGOS
# ============================================================

@st.cache_data(ttl=600)
def obtener_catalogos():

    data = obtener_datos(
        "resumen-mes-filtros"
    )

    if not data:
        return {
            "anios": [],
            "meses": [],
            "dias": [],
            "categorias": [],
            "clusters": [],
            "proveedores": [],
            "clientes": [],
            "sucursales": [],
        }

    return data


catalogos = obtener_catalogos()


anios = [
    int(x)
    for x in catalogos.get("anios", [])
    if x is not None
]

meses = [
    int(x)
    for x in catalogos.get("meses", [])
    if x is not None
]

dias = [
    int(x)
    for x in catalogos.get("dias", [])
    if x is not None
]

categorias = [
    str(x)
    for x in catalogos.get("categorias", [])
    if str(x).strip()
]

clusters = [
    str(x)
    for x in catalogos.get("clusters", [])
    if str(x).strip()
]

proveedores = [
    str(x)
    for x in catalogos.get("proveedores", [])
    if str(x).strip()
]

clientes = [
    str(x)
    for x in catalogos.get("clientes", [])
    if str(x).strip()
]

sucursales = [
    str(x)
    for x in catalogos.get("sucursales", [])
    if str(x).strip()
]


if not anios:
    st.error(
        "No se encontraron años en ia_tarjetas_club."
    )
    st.stop()


# ============================================================
# SESSION STATE
# ============================================================

if "rm_anio" not in st.session_state:
    st.session_state.rm_anio = anios[0]

if "rm_meses" not in st.session_state:
    st.session_state.rm_meses = (meses[0], meses[0]) if meses else (1, 1)

if "rm_dia" not in st.session_state:
    st.session_state.rm_dia = None

if "rm_categoria" not in st.session_state:
    st.session_state.rm_categoria = []

if "rm_cluster" not in st.session_state:
    st.session_state.rm_cluster = []

if "rm_proveedor" not in st.session_state:
    st.session_state.rm_proveedor = []

if "rm_cliente" not in st.session_state:
    st.session_state.rm_cliente = []

if "rm_sucursal" not in st.session_state:
    st.session_state.rm_sucursal = []

if "rm_consultado" not in st.session_state:
    st.session_state.rm_consultado = False


# ============================================================
# HELPERS
# ============================================================

def lista_parametro(valores):
    if not valores:
        return None
    return ",".join(str(x) for x in valores)


def construir_params():
    params = {
        "anio": st.session_state.rm_anio,
    }

    if st.session_state.rm_meses:
        mes_inicio, mes_fin = st.session_state.rm_meses
        params["mes_inicio"] = mes_inicio
        params["mes_fin"] = mes_fin

    if st.session_state.rm_dia is not None:
        params["dia"] = st.session_state.rm_dia

    categoria = lista_parametro(st.session_state.rm_categoria)
    if categoria:
        params["categoria"] = categoria

    cluster = lista_parametro(st.session_state.rm_cluster)
    if cluster:
        params["cluster"] = cluster

    proveedor = lista_parametro(st.session_state.rm_proveedor)
    if proveedor:
        params["proveedor"] = proveedor

    cliente = lista_parametro(st.session_state.rm_cliente)
    if cliente:
        params["nombre_cliente"] = cliente

    sucursal = lista_parametro(st.session_state.rm_sucursal)
    if sucursal:
        params["sucursal"] = sucursal

    return params


def limpiar_filtros():
    st.session_state.rm_anio = anios[0]
    st.session_state.rm_meses = (meses[0], meses[0]) if meses else (1, 1)
    st.session_state.rm_dia = None
    st.session_state.rm_categoria = []
    st.session_state.rm_cluster = []
    st.session_state.rm_proveedor = []
    st.session_state.rm_cliente = []
    st.session_state.rm_sucursal = []
    st.session_state.rm_consultado = False
    st.cache_data.clear()


def procesar_formatos_visuales(df):
    df = df.copy()

    if "Ticket_Promedio" in df.columns:
        df["Tickets Promedios"] = df["Ticket_Promedio"].apply(
            lambda x: f"{x:,.2f} ⭐" if pd.notnull(x) and x > 0 else "0.00 ⭐"
        )
        df.drop(columns=["Ticket_Promedio"], inplace=True)

    if "Pct_Dif_VTA" in df.columns:
        df["%Dif VTA"] = df["Pct_Dif_VTA"].apply(
            lambda x: f"{x * 100:.2f}% ⬇" if pd.notnull(x) and x < 0 else f"{x * 100:.2f}% ⬆" if pd.notnull(x) else "0.00%"
        )
        df.drop(columns=["Pct_Dif_VTA"], inplace=True)

    if "Pct_Dif_Abono" in df.columns:
        df["%Dif Abono"] = df["Pct_Dif_Abono"].apply(
            lambda x: f"{x * 100:.2f}% ⬇" if pd.notnull(x) and x < 0 else f"{x * 100:.2f}% ⬆" if pd.notnull(x) else "0.00%"
        )
        df.drop(columns=["Pct_Dif_Abono"], inplace=True)

    return df


def render_kpi_card(icon, titulo, valor, label, delta=None):
    if delta is not None:
        if isinstance(delta, (int, float)):
            clase = "kpi-badge-positive" if delta >= 0 else "kpi-badge-negative"
            simbolo = "▲" if delta >= 0 else "▼"
            badge_html = f'<div class="{clase}">{simbolo} {delta:.2%}</div>'
        else:
            badge_html = f'<div class="kpi-badge-info">{delta}</div>'
    else:
        badge_html = '<div style="visibility: hidden; font-size: 10px; padding: 2px 6px; margin-top: 4px;">-</div>'

    card_html = f"""
    <div class="kpi-card">
        <div class="kpi-top">
            <div class="kpi-header">
                <span class="kpi-icon">{icon}</span>
                <span style="overflow:hidden; text-overflow:ellipsis; white-space:nowrap;">{titulo}</span>
            </div>
            <div class="kpi-value">{valor}</div>
            {badge_html}
        </div>
        <div class="kpi-label" title="{label}">{label}</div>
    </div>
    """
    st.markdown(card_html, unsafe_allow_html=True)


def render_tabla_html(df, labels=None, formats=None, max_height=None):
    """
    Renderiza un DataFrame como tabla HTML con encabezados
    de color de marca (#FF5100) y texto blanco.

    Parámetros
    ----------
    df : pd.DataFrame
        DataFrame ya filtrado/ordenado con las columnas a mostrar,
        en el orden en que deben aparecer.
    labels : dict, opcional
        Mapeo {columna_original: nombre_a_mostrar}.
        Las columnas no incluidas conservan su nombre original.
    formats : dict, opcional
        Mapeo {columna_original: formato_str}, por ejemplo:
        {"VTA_Actual": "${:,.0f}", "Margen_Frontal": "{:.2f}%"}
        Se aplica ANTES de renombrar columnas.
    max_height : int, opcional
        Alto máximo en px; si se define, la tabla hace scroll vertical.
    """
    labels = labels or {}
    formats = formats or {}

    df_out = df.copy()

    for col, fmt in formats.items():
        if col in df_out.columns:
            df_out[col] = df_out[col].apply(
                lambda x: fmt.format(x) if pd.notnull(x) else ""
            )

    df_out = df_out.rename(columns=labels)

    styler = (
        df_out.style
        .hide(axis="index")
        .set_table_attributes('class="tabla-generada"')
    )

    tabla_html = styler.to_html()

    estilo_alto = (
        f"max-height:{max_height}px;" if max_height else ""
    )

    st.markdown(
        f'<div class="tabla-html-wrapper" style="{estilo_alto}">{tabla_html}</div>',
        unsafe_allow_html=True,
    )


def generar_excel_bytes(df, labels=None, nombre_hoja="Datos"):
    """
    Convierte un DataFrame a bytes de un archivo .xlsx en memoria,
    conservando los valores numéricos originales (sin $, % ni texto)
    para que se puedan usar fórmulas directamente en Excel.

    Intenta usar 'xlsxwriter' (permite colorear el header fácilmente);
    si no está instalado, hace fallback a 'openpyxl' (viene con pandas
    en la mayoría de instalaciones) aplicando el mismo estilo.

    labels : dict opcional {columna_original: nombre_a_mostrar}
    """
    labels = labels or {}

    df_excel = df.copy().rename(columns=labels)

    buffer = io.BytesIO()

    try:
        import xlsxwriter  # noqa: F401
        motor = "xlsxwriter"
    except ImportError:
        motor = "openpyxl"

    if motor == "xlsxwriter":
        with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
            df_excel.to_excel(writer, index=False, sheet_name=nombre_hoja)

            workbook = writer.book
            worksheet = writer.sheets[nombre_hoja]

            formato_header = workbook.add_format(
                {
                    "bold": True,
                    "bg_color": COLOR_MARCA,
                    "font_color": "#FFFFFF",
                    "border": 1,
                }
            )

            for col_idx, col_name in enumerate(df_excel.columns):
                worksheet.write(0, col_idx, col_name, formato_header)
                ancho = max(12, min(35, len(str(col_name)) + 4))
                worksheet.set_column(col_idx, col_idx, ancho)

    else:
        from openpyxl.styles import Font, PatternFill

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df_excel.to_excel(writer, index=False, sheet_name=nombre_hoja)

            worksheet = writer.sheets[nombre_hoja]

            relleno = PatternFill(
                start_color="FFFF5100",
                end_color="FFFF5100",
                fill_type="solid",
            )
            fuente = Font(color="FFFFFFFF", bold=True)

            for col_idx, col_name in enumerate(df_excel.columns, start=1):
                celda = worksheet.cell(row=1, column=col_idx)
                celda.fill = relleno
                celda.font = fuente

                ancho = max(12, min(35, len(str(col_name)) + 4))
                letra_col = celda.column_letter
                worksheet.column_dimensions[letra_col].width = ancho

    buffer.seek(0)
    return buffer


def render_boton_descarga(df, nombre_archivo, labels=None, key=None):
    """
    Dibuja un botón de descarga a Excel para el DataFrame dado.
    """
    excel_bytes = generar_excel_bytes(df, labels=labels)

    st.download_button(
        label="⬇️ Descargar Excel",
        data=excel_bytes,
        file_name=f"{nombre_archivo}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
        key=key,
    )


# ============================================================
# KPI SUPERIORES
# ============================================================

def render_kpi_superiores():
    data_kpi = obtener_datos("resumen-mes-presentacion")

    if not data_kpi:
        return

    # Mapeo directo con las columnas de IA_KPI_CLUB_TABLA
    vTotalClub = float(data_kpi.get("TotalClub", 0) or 0)
    vVentaYTD = float(data_kpi.get("Venta_Year_to_day", 0) or 0)
    vCPYTD = float(data_kpi.get("CP_Year_to_day", 0) or 0)
    vAbono = float(data_kpi.get("Abono", 0) or 0)
    vMargenYTD = float(data_kpi.get("Margen_Year_to_day", 0) or 0)

    # 5 columnas alineadas a los nuevos datos
    c1, c2, c3, c4, c5 = st.columns(5)

    with c1:
        render_kpi_card(
            icon="storefront",
            titulo="Total Club",
            valor=f"${vTotalClub:,.2f}",
            label="Total Acumulado",
            delta="Club"
        )

    with c2:
        render_kpi_card(
            icon="credit_card",
            titulo="Venta YTD",
            valor=f"${vVentaYTD:,.2f}",
            label="Venta Year to Date",
            delta="YTD"
        )

    with c3:
        render_kpi_card(
            icon="shopping_basket",
            titulo="CP YTD",
            valor=f"${vCPYTD:,.2f}",
            label="Costo Producción/Real",
            delta="CP"
        )

    with c4:
        render_kpi_card(
            icon="payments",
            titulo="Abono",
            valor=f"{vAbono:.2f}%" if abs(vAbono) <= 100 else f"${vAbono:,.2f}",
            label="Abono Registrado",
            delta="Abono"
        )

    with c5:
        render_kpi_card(
            icon="trending_up",
            titulo="Margen YTD",
            valor=f"{vMargenYTD:.2f}%" if abs(vMargenYTD) > 1 else f"{vMargenYTD * 100:.2f}%",
            label="Margen Year to Date",
            delta="Margen"
        )

render_kpi_superiores()


# ============================================================
# TÍTULO
# ============================================================



# ============================================================
# FILTROS
# ============================================================
render_divider()
render_section_header(
    icon="filter_alt", 
    titulo="Filtros de Búsqueda", 
    subtitulo="Selecciona los parámetros para actualizar la información en pantalla"
)

st.caption("**Filtros Obligatorios**")
with st.container(border=True):
    col_ob1, col_ob2, col_ob3 = st.columns(3)

    with col_ob1:
        anio_actual = st.selectbox(
            "Año *",
            options=anios,
            index=(
                anios.index(st.session_state.rm_anio)
                if st.session_state.rm_anio in anios
                else 0
            ),
            key="rm_anio_selector",
        )

    with col_ob2:
        meses_actuales = st.select_slider(
            "Rango de Meses *",
            options=meses,
            value=st.session_state.rm_meses,
            format_func=lambda x: DICC_MESES.get(x, str(x)),
            key="rm_mes_range_selector",
        )

    with col_ob3:
        cluster_actual = st.multiselect(
            "Cluster *",
            options=clusters,
            default=[x for x in st.session_state.rm_cluster if x in clusters],
            placeholder="Selecciona cluster...",
            key="rm_cluster_selector",
        )


with st.expander("**Filtros Adicionales (Opcionales)**", expanded=False):

    col_op1, col_op2, col_op3 = st.columns(3)

    with col_op1:
        opciones_dia = [None] + dias
        dia_actual = st.selectbox(
            "Día",
            options=opciones_dia,
            format_func=lambda x: "Todos los días" if x is None else str(x),
            index=(
                opciones_dia.index(st.session_state.rm_dia)
                if st.session_state.rm_dia in opciones_dia
                else 0
            ),
            key="rm_dia_selector",
        )

    with col_op2:
        categoria_actual = st.multiselect(
            "Categoría",
            options=categorias,
            default=[x for x in st.session_state.rm_categoria if x in categorias],
            placeholder="Todas las categorías",
            key="rm_categoria_selector",
        )

    with col_op3:
        proveedor_actual = st.multiselect(
            "Proveedor",
            options=proveedores,
            default=[x for x in st.session_state.rm_proveedor if x in proveedores],
            placeholder="Todos los proveedores",
            key="rm_proveedor_selector",
        )

    st.write("")

    col_op4, col_op5 = st.columns(2)

    with col_op4:
        cliente_actual = st.multiselect(
            "Nombre de cliente",
            options=clientes,
            default=[x for x in st.session_state.rm_cliente if x in clientes],
            placeholder="Todos los clientes",
            key="rm_cliente_selector",
        )

    with col_op5:
        sucursal_actual = st.multiselect(
            "Nombre de sucursal",
            options=sucursales,
            default=[x for x in st.session_state.rm_sucursal if x in sucursales],
            placeholder="Todas las sucursales",
            key="rm_sucursal_selector",
        )


st.write("")

b1, b2, b3 = st.columns([1, 1, 4])

with b1:
    consultar = st.button(
        "Consultar",
        type="primary",
        use_container_width=True,
        key="rm_consultar",
    )

with b2:
    limpiar = st.button(
        "Limpiar",
        use_container_width=True,
        key="rm_limpiar",
    )


# ============================================================
# APLICAR FILTROS Y VALIDACIÓN
# ============================================================

if limpiar:
    limpiar_filtros()
    st.rerun()


if consultar:
    if anio_actual is None or meses_actuales is None or not cluster_actual:
        st.warning("Debes seleccionar **Año**, **Rango de Meses** y al menos un **Cluster**.")
        st.session_state.rm_consultado = False
        st.stop()
    else:
        st.session_state.rm_anio = int(anio_actual)
        st.session_state.rm_meses = meses_actuales
        st.session_state.rm_dia = int(dia_actual) if dia_actual is not None else None
        st.session_state.rm_categoria = categoria_actual
        st.session_state.rm_cluster = cluster_actual
        st.session_state.rm_proveedor = proveedor_actual
        st.session_state.rm_cliente = cliente_actual
        st.session_state.rm_sucursal = sucursal_actual

        st.session_state.rm_consultado = True
        st.cache_data.clear()
        st.rerun()


# ============================================================
# CONTEXTO
# ============================================================

params_filtros = construir_params()


# ============================================================
# INDICADOR DE FILTROS ACTIVOS
# ============================================================

if st.session_state.rm_consultado:
    filtros_activos = []
    filtros_activos.append(f"Año: {st.session_state.rm_anio}")

    if st.session_state.rm_meses:
        m_inicio, m_fin = st.session_state.rm_meses
        if m_inicio == m_fin:
            txt_mes = f"Mes: {DICC_MESES.get(m_inicio)}"
        else:
            txt_mes = f"Meses: {DICC_MESES.get(m_inicio)} - {DICC_MESES.get(m_fin)}"
        filtros_activos.append(txt_mes)

    if st.session_state.rm_dia is not None:
        filtros_activos.append(f"Día: {st.session_state.rm_dia}")
    if st.session_state.rm_categoria:
        filtros_activos.append("Categoría: " + ", ".join(st.session_state.rm_categoria))
    if st.session_state.rm_cluster:
        filtros_activos.append("Cluster: " + ", ".join(st.session_state.rm_cluster))
    if st.session_state.rm_proveedor:
        filtros_activos.append("Proveedor: " + ", ".join(st.session_state.rm_proveedor))
    if st.session_state.rm_cliente:
        filtros_activos.append("Cliente: " + ", ".join(st.session_state.rm_cliente))
    if st.session_state.rm_sucursal:
        filtros_activos.append("Sucursal: " + ", ".join(st.session_state.rm_sucursal))

    st.info(" | ".join(filtros_activos))


# ============================================================
# SI TODAVÍA NO CONSULTA
# ============================================================

if not st.session_state.rm_consultado:
    st.info("Selecciona los filtros y presiona **Consultar** para cargar el Resumen del Mes.")
    st.stop()


# ============================================================
# PESTAÑAS
# ============================================================

tab_clientes, tab_sucursal, tab_cluster = st.tabs(
    [
        "Resumen de Abono y Redención por cliente",
        "Resumen por Sucursal",
        "Resumen por Categoría / Cluster",
    ]
)


# ============================================================
# TAB CLIENTES
# ============================================================

with tab_clientes:
    render_section_header(
        icon="group", 
        titulo="Resumen de Abono y Redención por Cliente", 
        subtitulo="Detalle acumulado de abonos, redenciones y saldos por cliente"
    )

    data_matriz = obtener_datos(
        "resumen-mes-clientes-matriz",
        params_filtros,
    )

    if data_matriz:
        df_mat = pd.DataFrame(data_matriz)

        if df_mat.empty:
            st.info("No existen clientes para los filtros seleccionados.")
        else:
            # 1. Métricas Totales
            tot_clientes = df_mat["Nombre_Cliente"].nunique()
            tot_abono = pd.to_numeric(df_mat["Abono_Actual"], errors="coerce").fillna(0).sum()
            tot_redencion = pd.to_numeric(df_mat["Redencion_Actual"], errors="coerce").fillna(0).sum()
            max_bonif = pd.to_numeric(df_mat["Max_Bonificacion"], errors="coerce").fillna(0).max()

            k1, k2, k3, k4 = st.columns(4)

            with k1:
                render_kpi_card(
                    icon="group",
                    titulo="Clientes Encontrados",
                    valor=f"{tot_clientes:,}",
                    label="Total de Clientes",
                    delta="Clientes"
                )

            with k2:
                render_kpi_card(
                    icon="payments",
                    titulo="Abono Total",
                    valor=f"${tot_abono:,.2f}",
                    label="Monto Acumulado",
                    delta="Abono"
                )

            with k3:
                render_kpi_card(
                    icon="shopping_basket",
                    titulo="Redención Total",
                    valor=f"${tot_redencion:,.2f}",
                    label="Monto Redimido",
                    delta="Redención"
                )

            with k4:
                render_kpi_card(
                    icon="stars",
                    titulo="Máx. Bonificación",
                    valor=f"${max_bonif:,.2f}",
                    label="Límite Máximo",
                    delta="Bonificación"
                )

            st.write("")

            # 2. Búsqueda, Tabla y Gráfica lado a lado
            busqueda = st.text_input(
                "🔎 Buscar cliente",
                placeholder="Escribe nombre del cliente...",
                key="rm_busqueda_cliente",
            )

            if busqueda.strip():
                mascara = (
                    df_mat["Nombre_Cliente"]
                    .fillna("")
                    .str.contains(busqueda, case=False, na=False)
                )
                df_mat = df_mat[mascara]

            if df_mat.empty:
                st.warning("No se encontraron coincidencias.")
            else:
                col_tab1, col_tab2 = st.columns([6, 4])

                with col_tab1:
                    columnas_deseadas = [
                        "Nombre_Cliente",
                        "Tarjeta",
                        "Cluster",
                        "Abono_Actual",
                        "Redencion_Actual",
                        "Max_Bonificacion",
                        "Limite_Bonificacion",
                    ]

                    columnas_mostrar = [col for col in columnas_deseadas if col in df_mat.columns]

                    render_tabla_html(
                        df_mat[columnas_mostrar],
                        labels={
                            "Nombre_Cliente": "Cliente",
                            "Tarjeta": "Tarjeta",
                            "Cluster": "Cluster",
                            "Abono_Actual": "Abono Actual",
                            "Redencion_Actual": "Redención Actual",
                            "Max_Bonificacion": "Max Bonificación",
                            "Limite_Bonificacion": "Límite Bonificación",
                        },
                        formats={
                            "Abono_Actual": "${:,.2f}",
                            "Redencion_Actual": "${:,.2f}",
                            "Max_Bonificacion": "${:,.2f}",
                            "Limite_Bonificacion": "${:,.2f}",
                        },
                        max_height=460,
                    )

                    render_boton_descarga(
                        df_mat[columnas_mostrar],
                        nombre_archivo="resumen_clientes",
                        labels={
                            "Nombre_Cliente": "Cliente",
                            "Tarjeta": "Tarjeta",
                            "Cluster": "Cluster",
                            "Abono_Actual": "Abono Actual",
                            "Redencion_Actual": "Redención Actual",
                            "Max_Bonificacion": "Max Bonificación",
                            "Limite_Bonificacion": "Límite Bonificación",
                        },
                        key="descarga_clientes",
                    )

                with col_tab2:
                    top_clientes = (
                        df_mat.copy()
                        .assign(Abono_Num=lambda x: pd.to_numeric(x["Abono_Actual"], errors="coerce").fillna(0))
                        .assign(Redencion_Num=lambda x: pd.to_numeric(x["Redencion_Actual"], errors="coerce").fillna(0))
                        .sort_values(by="Abono_Num", ascending=False)
                        .head(10)
                    )

                    if not top_clientes.empty:
                        st.markdown("**Top 10 Clientes con Mayor Abono vs Redención**")
                        
                        df_top_melted = top_clientes.melt(
                            id_vars=["Nombre_Cliente"], 
                            value_vars=["Abono_Num", "Redencion_Num"],
                            var_name="Métrica", 
                            value_name="Monto"
                        )
                        df_top_melted["Métrica"] = df_top_melted["Métrica"].map({"Abono_Num": "Abono", "Redencion_Num": "Redención"})

                        fig_top = px.bar(
                            df_top_melted,
                            x="Nombre_Cliente",
                            y="Monto",
                            color="Métrica",
                            barmode="group",
                            color_discrete_map={"Abono": "#FF5100", "Redención": "#0B47D4"}
                        )
                        fig_top.update_traces(
                            hovertemplate="<b>%{x}</b><br>%{data.name}: $%{y:,.2f}<extra></extra>"
                        )
                        fig_top.update_layout(
                            margin=dict(l=10, r=10, t=10, b=10),
                            height=400,
                            xaxis_title=None,
                            yaxis_title=None,
                            legend_title_text=None
                        )
                        st.plotly_chart(fig_top, use_container_width=True)

    else:
        st.warning("No se pudo obtener información de clientes.")


# ============================================================
# TAB SUCURSAL
# ============================================================

with tab_sucursal:
    render_section_header(
        icon="storefront", 
        titulo="Resumen por Sucursal", 
        subtitulo="Métricas operativas y de ventas distribuidas por sucursal"
    )

    with st.spinner("Consultando sucursales..."):
        data_suc = obtener_datos(
            "resumen-mes-sucursal",
            params_filtros,
        )

    if data_suc:
        df_suc = pd.DataFrame(data_suc)

        if df_suc.empty:
            st.info("No existen datos de sucursales para los filtros seleccionados.")
        else:
            col_suc1, col_suc2 = st.columns([6, 4])

            with col_suc1:
                df_suc_tabla = procesar_formatos_visuales(df_suc)

                render_tabla_html(
                    df_suc_tabla,
                    labels={
                        "sucursal": "Sucursal",
                        "Share_Vta": "Share Vta",
                        "Tarjetas_Activas": "Tarjetas Activas",
                        "Clientes_App_Web": "Clientes App Web",
                        "Num_SKU": "#SKU",
                        "VTA_Actual": "VTA Actual",
                        "VTA_AA": "VTA AA",
                        "Abono_Actual": "Abono Actual",
                        "Abono_AA": "Abono AA",
                        "Margen_Frontal": "Margen Frontal",
                        "Margen_Frontal_Bonificacion": "Margen C/B",
                    },
                    formats={
                        "Share_Vta": "{:.2f}%",
                        "VTA_Actual": "${:,.0f}",
                        "VTA_AA": "${:,.0f}",
                        "Abono_Actual": "${:,.0f}",
                        "Abono_AA": "${:,.0f}",
                        "Margen_Frontal": "{:.2f}%",
                        "Margen_Frontal_Bonificacion": "{:.2f}%",
                    },
                    max_height=460,
                )

                render_boton_descarga(
                    df_suc_tabla,
                    nombre_archivo="resumen_sucursal",
                    labels={
                        "sucursal": "Sucursal",
                        "Share_Vta": "Share Vta",
                        "Tarjetas_Activas": "Tarjetas Activas",
                        "Clientes_App_Web": "Clientes App Web",
                        "Num_SKU": "#SKU",
                        "VTA_Actual": "VTA Actual",
                        "VTA_AA": "VTA AA",
                        "Abono_Actual": "Abono Actual",
                        "Abono_AA": "Abono AA",
                        "Margen_Frontal": "Margen Frontal",
                        "Margen_Frontal_Bonificacion": "Margen C/B",
                    },
                    key="descarga_sucursal",
                )

            with col_suc2:
                # Filtrar fila de Total y obtener Top 10 por VTA_Actual
                df_suc_grafica = (
                    df_suc[df_suc["sucursal"] != "Total"].copy()
                    if "sucursal" in df_suc.columns
                    else df_suc.copy()
                )

                if "VTA_Actual" in df_suc_grafica.columns:
                    df_suc_grafica["VTA_Actual_Num"] = pd.to_numeric(
                        df_suc_grafica["VTA_Actual"], errors="coerce"
                    ).fillna(0)

                    top_10_suc = df_suc_grafica.sort_values(
                        by="VTA_Actual_Num", ascending=False
                    ).head(10)

                    st.markdown("**Top 10 Sucursales con Mayor Venta**")
                    fig_vta_suc = px.bar(
                        top_10_suc,
                        x="sucursal",
                        y="VTA_Actual_Num",
                        color_discrete_sequence=["#FF5100"],
                    )
                    fig_vta_suc.update_traces(
                        hovertemplate="<b>%{x}</b><br>VTA Actual: $%{y:,.0f}<extra></extra>"
                    )
                    fig_vta_suc.update_layout(
                        margin=dict(l=10, r=10, t=10, b=10),
                        height=400,
                        xaxis_title=None,
                        yaxis_title=None,
                    )
                    st.plotly_chart(fig_vta_suc, use_container_width=True)

    else:
        st.warning("No se pudo obtener el resumen de sucursales.")


# ============================================================
# TAB CATEGORÍA / CLUSTER
# ============================================================

with tab_cluster:
    render_section_header(
        icon="category", 
        titulo="Resumen por Categoría / Cluster", 
        subtitulo="Desglose de rendimiento e ingresos agrupados por categoría"
    )

    with st.spinner("Consultando categorías..."):
        data_cat = obtener_datos(
            "resumen-mes-categoria",
            params_filtros,
        )

    if data_cat:
        df_cat = pd.DataFrame(data_cat)

        if df_cat.empty:
            st.info("No existen categorías para los filtros seleccionados.")
        else:
            col_cat1, col_cat2 = st.columns([6, 4])

            with col_cat1:
                df_cat_tabla = procesar_formatos_visuales(df_cat)

                render_tabla_html(
                    df_cat_tabla,
                    labels={
                        "Categoria_GB": "Categoría",
                        "Share_Cliente": "Share Cliente",
                        "Tarjetas_Activas": "Tarjetas Activas",
                        "Clientes_App_Web": "Clientes App Web",
                        "Num_SKU": "#SKU",
                        "VTA_Actual": "VTA Actual",
                        "VTA_AA": "VTA AA",
                        "Abono_Actual": "Abono Actual",
                        "Abono_AA": "Abono AA",
                        "Margen_Frontal": "Margen Frontal",
                    },
                    formats={
                        "Share_Cliente": "{:.2f}%",
                        "VTA_Actual": "${:,.0f}",
                        "VTA_AA": "${:,.0f}",
                        "Abono_Actual": "${:,.0f}",
                        "Abono_AA": "${:,.0f}",
                        "Margen_Frontal": "{:.2f}%",
                    },
                    max_height=460,
                )

                render_boton_descarga(
                    df_cat_tabla,
                    nombre_archivo="resumen_categoria",
                    labels={
                        "Categoria_GB": "Categoría",
                        "Share_Cliente": "Share Cliente",
                        "Tarjetas_Activas": "Tarjetas Activas",
                        "Clientes_App_Web": "Clientes App Web",
                        "Num_SKU": "#SKU",
                        "VTA_Actual": "VTA Actual",
                        "VTA_AA": "VTA AA",
                        "Abono_Actual": "Abono Actual",
                        "Abono_AA": "Abono AA",
                        "Margen_Frontal": "Margen Frontal",
                    },
                    key="descarga_categoria",
                )

            with col_cat2:
                # Filtrar fila Total y obtener Top 15 por Share/Venta
                df_cat_grafica = (
                    df_cat[df_cat["Categoria_GB"] != "Total"].copy()
                    if "Categoria_GB" in df_cat.columns
                    else df_cat.copy()
                )

                if "Share_Cliente" in df_cat_grafica.columns:
                    df_cat_grafica["Share_Num"] = pd.to_numeric(
                        df_cat_grafica["Share_Cliente"], errors="coerce"
                    ).fillna(0)

                    top_15_cat = df_cat_grafica.sort_values(
                        by="Share_Num", ascending=False
                    ).head(15)

                    st.markdown("**Top 15 Categorías por Participación**")
                    fig_pie_cat = px.pie(
                        top_15_cat,
                        names="Categoria_GB",
                        values="Share_Num",
                        hole=0.4,
                        color_discrete_sequence=px.colors.qualitative.Set2,
                    )
                    fig_pie_cat.update_traces(
                        hovertemplate="<b>%{label}</b><br>Share: %{value:.2f}%<extra></extra>"
                    )
                    fig_pie_cat.update_layout(
                        margin=dict(l=10, r=10, t=10, b=10),
                        height=400,
                    )
                    st.plotly_chart(fig_pie_cat, use_container_width=True)

    else:
        st.warning("No se pudo obtener el resumen por categoría.")